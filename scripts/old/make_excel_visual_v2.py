import json
from pathlib import Path
from statistics import mean

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl import load_workbook
from PIL import Image as PILImage

# ============================================================================
# Configure paths
# ============================================================================
BASE_DIR = Path("/Users/hy/PycharmProjects/PythonProject/Defect_VLM_Project/logs/consistency_v3_rephrase_viewpoint_sbert_first/20260426_233024/candidates")
OUTPUT_XLSX = BASE_DIR / "candidate_image_prompt_review_sheet_visual_v2.xlsx"

CANDIDATE_RECORD_FILES = {
    "rephrase_low": BASE_DIR / "rephrase_sbert_mean_score_low_records.json",
    "rephrase_high": BASE_DIR / "rephrase_sbert_mean_score_high_records.json",
    "rephrase_random": BASE_DIR / "rephrase_sbert_mean_score_random_records.json",
    "viewpoint_low": BASE_DIR / "viewpoint_sbert_mean_score_low_records.json",
    "viewpoint_high": BASE_DIR / "viewpoint_sbert_mean_score_high_records.json",
    "viewpoint_random": BASE_DIR / "viewpoint_sbert_mean_score_random_records.json",
}

THUMB_DIR = BASE_DIR / "_thumbs_v2"
THUMB_DIR.mkdir(exist_ok=True)

GROUP_ORDER = [
    "rephrase_low",
    "rephrase_high",
    "rephrase_random",
    "viewpoint_low",
    "viewpoint_high",
    "viewpoint_random",
]

PROMPT_TYPE_FILL = {
    "rephrase": "DDEBF7",
    "viewpoint": "FCE4D6",
}

CANDIDATE_FILL = {
    "low": "FDE9E7",
    "high": "E2F0D9",
    "random": "EDEDED",
}

THIN = Side(style="thin", color="D9D9D9")
MEDIUM = Side(style="medium", color="A6A6A6")
BORDER_THIN = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
BORDER_MEDIUM = Border(top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=MEDIUM)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
META_FILL = PatternFill("solid", fgColor="F7F7F7")
PROMPT_FILL = PatternFill("solid", fgColor="FFF2CC")
RESPONSE_FILL = PatternFill("solid", fgColor="F2F2F2")
NOTE_FILL = PatternFill("solid", fgColor="FFF9E6")


# ============================================================================
# Utilities
# ============================================================================
def load_records(json_path: Path):
    if not json_path.exists():
        print(f"missing: {json_path}")
        return []
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def make_thumbnail(image_path, size=(180, 180)):
    image_path = Path(image_path)
    thumb_path = THUMB_DIR / image_path.name
    try:
        img = PILImage.open(image_path).convert("RGB")
        img.thumbnail(size)
        img.save(thumb_path)
        return thumb_path
    except Exception as e:
        print(f"thumbnail error: {image_path} | {e}")
        return None


def get_pairwise_min_max(record):
    pairs = record.get("sbert_pairwise", [])
    if not pairs:
        return None, None
    scores = [p["score"] for p in pairs]
    return min(scores), max(scores)


def parse_group_name(group_name: str):
    prompt_type, candidate_type = group_name.split("_", 1)
    return prompt_type, candidate_type


def flatten_records():
    rows = []
    for group_name in GROUP_ORDER:
        json_path = CANDIDATE_RECORD_FILES[group_name]
        records = load_records(json_path)
        prompt_type, candidate_type = parse_group_name(group_name)
        for idx, record in enumerate(records, start=1):
            image_path = Path(record["image_path"])
            pair_min, pair_max = get_pairwise_min_max(record)
            rows.append(
                {
                    "group_name": group_name,
                    "group_order": GROUP_ORDER.index(group_name),
                    "prompt_type": prompt_type,
                    "candidate_type": candidate_type,
                    "local_index": idx,
                    "class_name": record.get("class_name", image_path.parent.name),
                    "image_name": image_path.name,
                    "image_path": str(image_path),
                    "record": record,
                    "sbert_mean": record.get("sbert_mean_score"),
                    "pair_min": pair_min,
                    "pair_max": pair_max,
                    "prompt_records": record.get("prompt_records", []),
                }
            )
    return rows


def style_cell(cell, *, fill=None, font=None, alignment=None, border=None, number_format=None):
    if fill is not None:
        cell.fill = fill
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def fill_merged_range(ws, cell_range, fill=None, border=None, font=None, alignment=None):
    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            style_cell(cell, fill=fill, border=border, font=font, alignment=alignment)


def auto_wrap_and_border(ws, start_row, end_row, start_col, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(r, c).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(r, c).border = BORDER_THIN


# ============================================================================
# Sheet 1: Dashboard
# ============================================================================
def build_dashboard(wb, rows):
    ws = wb.active
    ws.title = "dashboard"

    ws["A1"] = "Candidate Review Dashboard"
    style_cell(
        ws["A1"],
        font=Font(size=16, bold=True, color="FFFFFF"),
        fill=HEADER_FILL,
        alignment=Alignment(horizontal="center", vertical="center"),
    )
    ws.merge_cells("A1:H1")
    fill_merged_range(ws, "A1:H1", fill=HEADER_FILL, border=BORDER_MEDIUM)

    info_rows = [
        ("Purpose", "한 장짜리 wide sheet 대신 summary + review_cards + sortable_table 구조로 재구성"),
        ("Main visual change", "prompt/response를 가로 병렬 대신 세로 스택으로 배치해 읽기 부담을 줄임"),
        ("Color meaning", "파랑=rephrase, 주황=viewpoint, 빨강=low, 초록=high, 회색=random"),
        ("Reviewer action", "review_cards에서 이미지/응답 확인 후 점수와 Final Use for AL? 입력"),
    ]

    start = 3
    for i, (k, v) in enumerate(info_rows, start=start):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
        style_cell(ws[f"A{i}"], font=Font(bold=True), fill=SUBHEADER_FILL, border=BORDER_THIN)
        style_cell(ws[f"B{i}"], alignment=Alignment(wrap_text=True), border=BORDER_THIN)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)
        fill_merged_range(ws, f"B{i}:H{i}", border=BORDER_THIN)

    summary_header_row = 9
    headers = ["Group", "Prompt Type", "Candidate Type", "Count", "SBERT Mean", "SBERT Min", "SBERT Max", "Pairwise Min Mean"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(summary_header_row, col_idx, header)
        style_cell(cell, fill=HEADER_FILL, font=HEADER_FONT, alignment=Alignment(horizontal="center", vertical="center"), border=BORDER_MEDIUM)

    current_row = summary_header_row + 1
    for group_name in GROUP_ORDER:
        subset = [r for r in rows if r["group_name"] == group_name]
        prompt_type, candidate_type = parse_group_name(group_name)
        sbert_vals = [r["sbert_mean"] for r in subset if r["sbert_mean"] is not None]
        pair_min_vals = [r["pair_min"] for r in subset if r["pair_min"] is not None]

        values = [
            group_name,
            prompt_type,
            candidate_type,
            len(subset),
            mean(sbert_vals) if sbert_vals else None,
            min(sbert_vals) if sbert_vals else None,
            max(sbert_vals) if sbert_vals else None,
            mean(pair_min_vals) if pair_min_vals else None,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(current_row, col_idx, value)
            fill = PatternFill("solid", fgColor=PROMPT_TYPE_FILL[prompt_type]) if col_idx in (1, 2) else PatternFill("solid", fgColor=CANDIDATE_FILL[candidate_type]) if col_idx == 3 else None
            style_cell(cell, fill=fill, alignment=Alignment(horizontal="center", vertical="center"), border=BORDER_THIN)
            if col_idx >= 5:
                cell.number_format = "0.000"
        current_row += 1

    legend_row = current_row + 2
    ws[f"A{legend_row}"] = "Legend"
    style_cell(ws[f"A{legend_row}"], fill=HEADER_FILL, font=HEADER_FONT, border=BORDER_MEDIUM)
    legends = [
        ("rephrase", PROMPT_TYPE_FILL["rephrase"]),
        ("viewpoint", PROMPT_TYPE_FILL["viewpoint"]),
        ("low", CANDIDATE_FILL["low"]),
        ("high", CANDIDATE_FILL["high"]),
        ("random", CANDIDATE_FILL["random"]),
    ]
    for i, (label, color) in enumerate(legends, start=legend_row + 1):
        ws[f"A{i}"] = label
        style_cell(ws[f"A{i}"], fill=PatternFill("solid", fgColor=color), border=BORDER_THIN)
        ws[f"B{i}"] = f"{label} group"
        style_cell(ws[f"B{i}"], border=BORDER_THIN)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 16
    ws.freeze_panes = "A9"

    if current_row > summary_header_row + 1:
        ws.conditional_formatting.add(
            f"E{summary_header_row + 1}:G{current_row - 1}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            )
        )


# ============================================================================
# Sheet 2: Review Cards (vertical)
# ============================================================================
def build_review_cards(wb, rows):
    ws = wb.create_sheet("review_cards")

    columns = {
        "A": 6,
        "B": 20,
        "C": 16,
        "D": 14,
        "E": 14,
        "F": 16,
        "G": 24,
        "H": 10,
        "I": 10,
        "J": 10,
        "K": 8,
        "L": 36,
        "M": 52,
        "N": 12,
        "O": 12,
        "P": 12,
        "Q": 12,
        "R": 12,
        "S": 16,
        "T": 32,
    }
    for col, width in columns.items():
        ws.column_dimensions[col].width = width

    headers = [
        "No", "Thumbnail", "Group", "Prompt Type", "Candidate Type", "Class", "Image Name",
        "SBERT", "Pair Min", "Pair Max", "P#", "Prompt", "Response",
        "Semantic", "Spatial", "Attribute", "Domain", "Total", "Final Use", "Manual Note"
    ]
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(1, i, header)
        style_cell(cell, fill=HEADER_FILL, font=HEADER_FONT, alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BORDER_MEDIUM)

    score_dv = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True)
    final_dv = DataValidation(type="list", formula1='"YES,NO,MAYBE"', allow_blank=True)
    ws.add_data_validation(score_dv)
    ws.add_data_validation(final_dv)

    row_ptr = 2
    overall_no = 1

    for group_name in GROUP_ORDER:
        group_rows = [r for r in rows if r["group_name"] == group_name]
        if not group_rows:
            continue

        prompt_type, candidate_type = parse_group_name(group_name)
        band_fill = PatternFill("solid", fgColor=PROMPT_TYPE_FILL[prompt_type])
        candidate_fill = PatternFill("solid", fgColor=CANDIDATE_FILL[candidate_type])

        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=20)
        title_cell = ws.cell(row_ptr, 1, f"{group_name}  |  {prompt_type} / {candidate_type}  |  {len(group_rows)} samples")
        style_cell(title_cell, fill=band_fill, font=Font(bold=True), alignment=Alignment(horizontal="left", vertical="center"), border=BORDER_MEDIUM)
        fill_merged_range(ws, f"A{row_ptr}:T{row_ptr}", fill=band_fill, border=BORDER_MEDIUM)
        row_ptr += 1

        for row in group_rows:
            record = row["record"]
            prompts = row["prompt_records"]
            block_start = row_ptr
            block_end = row_ptr + max(5, len(prompts)) - 1

            merge_cols = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 18, 19, 20]
            for col in merge_cols:
                ws.merge_cells(start_row=block_start, start_column=col, end_row=block_end, end_column=col)

            ws.cell(block_start, 1, overall_no)
            ws.cell(block_start, 3, row["group_name"])
            ws.cell(block_start, 4, row["prompt_type"])
            ws.cell(block_start, 5, row["candidate_type"])
            ws.cell(block_start, 6, row["class_name"])
            ws.cell(block_start, 7, row["image_name"])
            ws.cell(block_start, 8, row["sbert_mean"])
            ws.cell(block_start, 9, row["pair_min"])
            ws.cell(block_start, 10, row["pair_max"])
            ws.cell(block_start, 18, f"=SUM(N{block_start}:Q{block_start})")
            ws.cell(block_start, 19, "")
            ws.cell(block_start, 20, "")

            for col in [1, 3, 4, 5, 6, 7, 8, 9, 10, 18, 19, 20]:
                style_cell(
                    ws.cell(block_start, col),
                    fill=META_FILL,
                    alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                    border=BORDER_THIN,
                )

            fill_merged_range(ws, f"C{block_start}:C{block_end}", fill=band_fill, border=BORDER_THIN)
            fill_merged_range(ws, f"D{block_start}:D{block_end}", fill=band_fill, border=BORDER_THIN)
            fill_merged_range(ws, f"E{block_start}:E{block_end}", fill=candidate_fill, border=BORDER_THIN)
            for cell_ref in [f"H{block_start}", f"I{block_start}", f"J{block_start}", f"R{block_start}", f"S{block_start}", f"T{block_start}"]:
                ws[cell_ref].fill = NOTE_FILL

            ws.cell(block_start, 7).hyperlink = row["image_path"]
            ws.cell(block_start, 7).style = "Hyperlink"
            ws.cell(block_start, 19, "MAYBE")

            score_dv.add(f"N{block_start}:Q{block_start}")
            final_dv.add(f"S{block_start}")

            for metric_col in [8, 9, 10]:
                ws.cell(block_start, metric_col).number_format = "0.000"

            for i in range(max(5, len(prompts))):
                current_row = block_start + i
                ws.row_dimensions[current_row].height = 34
                ws.cell(current_row, 11, i + 1)
                style_cell(ws.cell(current_row, 11), fill=SUBHEADER_FILL, alignment=Alignment(horizontal="center", vertical="center"), border=BORDER_THIN)

                prompt_text = prompts[i].get("prompt_text", "") if i < len(prompts) else ""
                response_text = prompts[i].get("response_text", "") if i < len(prompts) else ""
                ws.cell(current_row, 12, prompt_text)
                ws.cell(current_row, 13, response_text)
                style_cell(ws.cell(current_row, 12), fill=PROMPT_FILL, alignment=Alignment(vertical="center", wrap_text=True), border=BORDER_THIN)
                style_cell(ws.cell(current_row, 13), fill=RESPONSE_FILL, alignment=Alignment(vertical="center", wrap_text=True), border=BORDER_THIN)

            thumb_path = make_thumbnail(row["image_path"])
            if thumb_path:
                img = XLImage(str(thumb_path))
                img.anchor = f"B{block_start}"
                ws.add_image(img)

            for r in range(block_start, block_end + 1):
                for c in range(1, 21):
                    ws.cell(r, c).border = BORDER_THIN

            for col in range(1, 21):
                ws.cell(block_start, col).border = Border(top=MEDIUM, bottom=THIN, left=THIN, right=THIN)
                ws.cell(block_end, col).border = Border(top=THIN, bottom=MEDIUM, left=THIN, right=THIN)

            overall_no += 1
            row_ptr = block_end + 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:T{max(2, row_ptr - 1)}"
    ws.conditional_formatting.add(
        f"H2:H{max(2, row_ptr - 1)}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        )
    )
    ws.conditional_formatting.add(
        f"R2:R{max(2, row_ptr - 1)}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="num", mid_value=4, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
    )


# ============================================================================
# Sheet 3: Sortable Review Table
# ============================================================================
def build_review_table(wb, rows):
    ws = wb.create_sheet("review_table")
    headers = [
        "No", "Group", "Prompt Type", "Candidate Type", "Class", "Image Name", "Image Path",
        "SBERT Mean", "Pairwise Min", "Pairwise Max", "Prompt 1", "Response 1", "Prompt 2", "Response 2",
        "Prompt 3", "Response 3", "Prompt 4", "Response 4", "Prompt 5", "Response 5",
        "Semantic", "Spatial", "Attribute", "Domain", "Total", "Final Use", "Manual Note"
    ]
    for idx, header in enumerate(headers, start=1):
        ws.cell(1, idx, header)
        style_cell(ws.cell(1, idx), fill=HEADER_FILL, font=HEADER_FONT, alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BORDER_MEDIUM)

    score_dv = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True)
    final_dv = DataValidation(type="list", formula1='"YES,NO,MAYBE"', allow_blank=True)
    ws.add_data_validation(score_dv)
    ws.add_data_validation(final_dv)

    for row_idx, row in enumerate(rows, start=2):
        prompts = row["prompt_records"]
        values = [
            row_idx - 1,
            row["group_name"], row["prompt_type"], row["candidate_type"], row["class_name"], row["image_name"], row["image_path"],
            row["sbert_mean"], row["pair_min"], row["pair_max"],
        ]
        for i in range(5):
            values.append(prompts[i].get("prompt_text", "") if i < len(prompts) else "")
            values.append(prompts[i].get("response_text", "") if i < len(prompts) else "")
        values.extend([None, None, None, None, f"=SUM(U{row_idx}:X{row_idx})", "MAYBE", ""])

        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)
            ws.cell(row_idx, col_idx).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row_idx, col_idx).border = BORDER_THIN

        ws.cell(row_idx, 6).hyperlink = row["image_path"]
        ws.cell(row_idx, 6).style = "Hyperlink"

        score_dv.add(f"U{row_idx}:X{row_idx}")
        final_dv.add(f"Z{row_idx}")

        prompt_fill = PatternFill("solid", fgColor=PROMPT_TYPE_FILL[row["prompt_type"]])
        cand_fill = PatternFill("solid", fgColor=CANDIDATE_FILL[row["candidate_type"]])
        ws.cell(row_idx, 2).fill = prompt_fill
        ws.cell(row_idx, 3).fill = prompt_fill
        ws.cell(row_idx, 4).fill = cand_fill
        for col in [8, 9, 10, 25]:
            ws.cell(row_idx, col).number_format = "0.000"

    widths = {
        1: 6, 2: 18, 3: 12, 4: 12, 5: 16, 6: 24, 7: 48,
        8: 12, 9: 12, 10: 12,
        11: 26, 12: 36, 13: 26, 14: 36, 15: 26, 16: 36, 17: 26, 18: 36, 19: 26, 20: 36,
        21: 10, 22: 10, 23: 10, 24: 10, 25: 10, 26: 12, 27: 24,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(2, len(rows) + 1)}"
    ws.conditional_formatting.add(
        f"H2:H{max(2, len(rows) + 1)}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        )
    )
    ws.conditional_formatting.add(
        f"Y2:Y{max(2, len(rows) + 1)}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="num", mid_value=4, mid_color="FFEB84",
            end_type="max", end_color="F8696B",
        )
    )


# ============================================================================
# Sheet 4: Scoring guide
# ============================================================================
def build_scoring_guide(wb):
    ws = wb.create_sheet("scoring_guide")
    guide_rows = [
        ["Criterion", "0", "1", "2", "Interpretation"],
        ["Semantic Inconsistency", "Same defect meaning", "Minor variation", "Different defect types", "예: dark spot ↔ stain 은 1, dark spot ↔ crack 은 2"],
        ["Spatial/Direction Inconsistency", "Same location/direction", "Minor difference", "Contradictory location/direction", "예: vertical ↔ horizontal 은 2"],
        ["Attribute Inconsistency", "Same shape/texture", "Minor difference", "Different shape/texture", "예: circular spot ↔ long streak 은 2"],
        ["Domain Error", "No error", "Weak questionable term", "Clear wrong domain", "예: metal defect인데 fabric/textile 표현이 나오면 2"],
        ["Total Difficulty", "0-2 easy", "3-5 moderate", "6-8 difficult", "초기 AL 후보 검토용 난이도 기준"],
        ["Final Use for AL?", "YES", "MAYBE", "NO", "라벨링 우선순위에 따라 선택"],
    ]

    for row in guide_rows:
        ws.append(row)

    for cell in ws[1]:
        style_cell(cell, fill=HEADER_FILL, font=HEADER_FONT, alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BORDER_MEDIUM)
    for r in range(2, len(guide_rows) + 1):
        for c in range(1, 6):
            style_cell(ws.cell(r, c), border=BORDER_THIN, alignment=Alignment(vertical="center", wrap_text=True))
        ws.cell(r, 1).fill = SUBHEADER_FILL
        ws.cell(r, 1).font = Font(bold=True)

    widths = {1: 28, 2: 22, 3: 22, 4: 28, 5: 46}
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


# ============================================================================
# Post-processing helper (optional)
# ============================================================================
def polish_existing_workbook(existing_xlsx: Path, new_output_xlsx: Path):
    wb = load_workbook(existing_xlsx)
    if "candidate_review" not in wb.sheetnames:
        raise ValueError("candidate_review sheet not found")
    ws = wb["candidate_review"]
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    wb.save(new_output_xlsx)
    print(f"polished existing workbook saved: {new_output_xlsx}")


# ============================================================================
# Main
# ============================================================================
def build_workbook():
    rows = flatten_records()
    wb = Workbook()
    build_dashboard(wb, rows)
    build_review_cards(wb, rows)
    build_review_table(wb, rows)
    build_scoring_guide(wb)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(OUTPUT_XLSX)
    print(f"saved: {OUTPUT_XLSX}")


if __name__ == "__main__":
    build_workbook()
