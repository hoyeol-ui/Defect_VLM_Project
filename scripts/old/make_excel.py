import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

BASE_DIR = Path("/Users/hy/PycharmProjects/PythonProject/Defect_VLM_Project/logs/consistency_v3_rephrase_viewpoint_sbert_first/20260426_233024/candidates")
OUTPUT_XLSX = BASE_DIR / "candidate_image_prompt_review_sheet.xlsx"

CANDIDATE_RECORD_FILES = {
    "rephrase_low": BASE_DIR / "rephrase_sbert_mean_score_low_records.json",
    "rephrase_high": BASE_DIR / "rephrase_sbert_mean_score_high_records.json",
    "rephrase_random": BASE_DIR / "rephrase_sbert_mean_score_random_records.json",
    "viewpoint_low": BASE_DIR / "viewpoint_sbert_mean_score_low_records.json",
    "viewpoint_high": BASE_DIR / "viewpoint_sbert_mean_score_high_records.json",
    "viewpoint_random": BASE_DIR / "viewpoint_sbert_mean_score_random_records.json",
}

THUMB_DIR = BASE_DIR / "_thumbs"
THUMB_DIR.mkdir(exist_ok=True)


def load_records(json_path):
    if not json_path.exists():
        print(f"missing: {json_path}")
        return []
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def make_thumbnail(image_path, size=(120, 120)):
    image_path = Path(image_path)
    thumb_path = THUMB_DIR / image_path.name

    try:
        img = PILImage.open(image_path).convert("RGB")
        img.thumbnail(size)
        img.save(thumb_path)
        return thumb_path
    except Exception as e:
        print(f"thumbnail error: {image_path} | {e}")
        return None


def get_pairwise_min_max(record):
    pairs = record.get("sbert_pairwise", [])
    if not pairs:
        return None, None

    scores = [p["score"] for p in pairs]
    return min(scores), max(scores)


def build_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = "candidate_review"

    headers = [
        "No",
        "Thumbnail",
        "Group",
        "Prompt Type",
        "Candidate Type",
        "Class",
        "Image Name",
        "SBERT Mean",
        "Pairwise Min",
        "Pairwise Max",
        "Image Path",
        "Prompt 1", "Response 1",
        "Prompt 2", "Response 2",
        "Prompt 3", "Response 3",
        "Prompt 4", "Response 4",
        "Prompt 5", "Response 5",
        "Semantic Inconsistency(0-2)",
        "Spatial/Direction Inconsistency(0-2)",
        "Attribute Inconsistency(0-2)",
        "Domain Error(0-2)",
        "Total Difficulty",
        "Manual Note",
        "Final Use for AL?",
    ]

    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="DDDDDD")

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    row_idx = 2
    no = 1

    for group_name, json_path in CANDIDATE_RECORD_FILES.items():
        records = load_records(json_path)

        for record in records:
            image_path = Path(record["image_path"])
            class_name = record.get("class_name", image_path.parent.name)
            image_name = image_path.name
            sbert_mean = record.get("sbert_mean_score")

            pair_min, pair_max = get_pairwise_min_max(record)

            parts = group_name.split("_")
            prompt_type = parts[0]
            candidate_type = parts[1]

            ws.cell(row=row_idx, column=1, value=no)
            ws.cell(row=row_idx, column=3, value=group_name)
            ws.cell(row=row_idx, column=4, value=prompt_type)
            ws.cell(row=row_idx, column=5, value=candidate_type)
            ws.cell(row=row_idx, column=6, value=class_name)
            ws.cell(row=row_idx, column=7, value=image_name)
            ws.cell(row=row_idx, column=8, value=sbert_mean)
            ws.cell(row=row_idx, column=9, value=pair_min)
            ws.cell(row=row_idx, column=10, value=pair_max)
            ws.cell(row=row_idx, column=11, value=str(image_path))

            prompt_records = record.get("prompt_records", [])

            start_col = 12
            for i in range(5):
                prompt_col = start_col + i * 2
                response_col = prompt_col + 1

                if i < len(prompt_records):
                    ws.cell(row=row_idx, column=prompt_col, value=prompt_records[i].get("prompt_text", ""))
                    ws.cell(row=row_idx, column=response_col, value=prompt_records[i].get("response_text", ""))
                else:
                    ws.cell(row=row_idx, column=prompt_col, value="")
                    ws.cell(row=row_idx, column=response_col, value="")

            # manual scoring columns
            semantic_col = 22
            spatial_col = 23
            attribute_col = 24
            domain_col = 25
            total_col = 26

            ws.cell(row=row_idx, column=total_col, value=f"=SUM(V{row_idx}:Y{row_idx})")

            thumb_path = make_thumbnail(image_path)
            if thumb_path:
                img = XLImage(str(thumb_path))
                img.anchor = f"B{row_idx}"
                ws.add_image(img)

            ws.row_dimensions[row_idx].height = 110

            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )
                ws.cell(row=row_idx, column=col).border = Border(
                    top=thin,
                    bottom=thin,
                    left=thin,
                    right=thin,
                )

            row_idx += 1
            no += 1

    widths = {
        1: 6,
        2: 18,
        3: 18,
        4: 14,
        5: 14,
        6: 18,
        7: 24,
        8: 12,
        9: 12,
        10: 12,
        11: 60,
        12: 36,
        13: 45,
        14: 36,
        15: 45,
        16: 36,
        17: 45,
        18: 36,
        19: 45,
        20: 36,
        21: 45,
        22: 22,
        23: 28,
        24: 24,
        25: 18,
        26: 16,
        27: 40,
        28: 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"

    ws2 = wb.create_sheet("scoring_guide")
    guide_rows = [
        ["Criterion", "0", "1", "2"],
        ["Semantic Inconsistency", "Same defect meaning", "Minor variation", "Different defect types"],
        ["Spatial/Direction Inconsistency", "Same location/direction", "Minor difference", "Contradictory location/direction"],
        ["Attribute Inconsistency", "Same shape/texture", "Minor difference", "Different shape/texture"],
        ["Domain Error", "No error", "Weak questionable term", "Clear wrong domain e.g. fabric/textile"],
        ["Total Difficulty", "0-2 easy", "3-5 moderate", "6-8 difficult"],
    ]

    for row in guide_rows:
        ws2.append(row)

    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for col in range(1, 5):
        ws2.column_dimensions[get_column_letter(col)].width = 30

    wb.save(OUTPUT_XLSX)
    print(f"saved: {OUTPUT_XLSX}")


if __name__ == "__main__":
    build_sheet()