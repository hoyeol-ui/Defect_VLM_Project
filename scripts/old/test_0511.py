import csv
import itertools
import json
import random
import statistics
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import torch
from sentence_transformers import SentenceTransformer
from transformers import AutoProcessor, Qwen2VLForConditionalGeneration
from qwen_vl_utils import process_vision_info

try:
    from bert_score import score as bertscore_score
except ImportError:
    bertscore_score = None


# =============================================================================
# [1] 전역 설정
# =============================================================================

DEVICE = "mps" if torch.backends.mps.is_available() else "cpu"

MODEL_ID   = "Qwen/Qwen2-VL-2B-Instruct"
SBERT_ID   = "sentence-transformers/all-MiniLM-L6-v2"

DATE_TAG   = datetime.now().strftime("%Y%m%d")   # → 20260511
TIME_TAG   = datetime.now().strftime("%H%M%S")

# ── 사용할 데이터셋 선택 ─────────────────────────────────────────────────────
# "NEU_DET" | "KOLEKTOR" | "MVTEC" 중 하나 또는 리스트로 지정
ACTIVE_DATASETS: List[str] = ["NEU_DET", "KOLEKTOR", "MVTEC"]

# ── 데이터셋별 절대 경로 ─────────────────────────────────────────────────────
# 참고: 실제 환경에 맞춰 수정하세요.
DATASET_PATHS: Dict[str, Path] = {
    # NEU-DET: 클래스별 하위 폴더 (crazing / inclusion / patches /
    #          pitted_surface / rolled-in_scale / scratches)
    "NEU_DET": Path(
        "/Users/hy/PycharmProjects/PythonProject"
        "/Defect_VLM_Project/data/NEU-DET/train/images"
    ),
    # KolektorSDD2-DatasetNinja: train/img 아래 이미지 파일
    # 어노테이션은 train/ann 폴더 (JSON 형식)
    "KOLEKTOR": Path(
        "/Users/hy/PycharmProjects/PythonProject"
        "/Defect_VLM_Project/data/Kolektor/kolektorsdd2-DatasetNinja/train/img"
    ),
    # MVTec AD: 카테고리별 폴더 → train/good, test/<defect_type>
    "MVTEC": Path(
        "/Users/hy/PycharmProjects/PythonProject"
        "/Defect_VLM_Project/data/MVTec/archive"
    ),
}

# ── MVTec 사용 카테고리 (None이면 전체 사용) ─────────────────────────────────
# 권장: 금속/질감 계열 우선
MVTEC_USE_CATEGORIES: Optional[List[str]] = [
    "metal_nut",
    "screw",
    "capsule",
    "pill",
    "transistor",
]

# ── MVTec test split 사용 (defect 유형별 폴더 존재) ─────────────────────────
MVTEC_USE_TEST_SPLIT: bool = True   # True → test/<defect_type>, False → train/good

# ── 샘플링 설정 ──────────────────────────────────────────────────────────────
SAMPLES_PER_CLASS  = 5   # 클래스/카테고리당 이미지 수 (초기 탐색용으로 작게 설정)
SEED               = 42

# ── 생성 파라미터 ─────────────────────────────────────────────────────────────
DO_SAMPLE       = True
TEMPERATURE     = 0.7
TOP_P           = 0.9
MAX_NEW_TOKENS  = 64

# ── Active Learning candidate 수 ─────────────────────────────────────────────
LOW_K    = 10
HIGH_K   = 10
RANDOM_K = 10

# ── BERTScore 설정 ────────────────────────────────────────────────────────────
USE_BERTSCORE_FULL        = False
USE_BERTSCORE_CANDIDATES  = False

# ── 출력 루트 경로 ────────────────────────────────────────────────────────────
OUTPUT_ROOT = Path(
    "/Users/hy/PycharmProjects/PythonProject/Defect_VLM_Project/logs"
)

EXPERIMENT_NAME = f"consistency_multi_dataset_{DATE_TAG}"

# ── Prompt 그룹 ───────────────────────────────────────────────────────────────
REPHRASE_PROMPTS = [
    "Describe the visible defect on this surface in one short sentence.",
    "In one short sentence, describe the defect visible on this surface.",
    "Briefly describe the visible defect on the surface in one sentence.",
    "Provide a one-sentence description of the visible defect on this surface.",
    "State in one short sentence what defect is visible on this surface.",
]

VIEWPOINT_PROMPTS = [
    "Describe the defect by focusing on its location and affected area in one short sentence.",
    "Describe the defect by focusing on its shape or structural pattern in one short sentence.",
    "Describe the defect by focusing on its texture, surface condition, or appearance in one short sentence.",
    "Describe the defect by focusing on whether the damaged region is localized or spread out in one short sentence.",
    "Describe the defect by focusing on the overall visual impression of the damaged surface in one short sentence.",
]

# ── 메서드 출처 메타 ──────────────────────────────────────────────────────────
METHOD_UPSTREAM_SOURCES = {
    "selfcheckgpt": {
        "official_repo": "https://github.com/potsawee/selfcheckgpt",
        "paper": "Manakul et al., SelfCheckGPT, arXiv:2303.08896",
        "adaptation_note": (
            "Multi-response self-consistency 개념을 VLM defect description에 적용."
        ),
    },
    "semantic_uncertainty": {
        "official_repo": "https://github.com/lorenzkuhn/semantic_uncertainty",
        "paper": "Kuhn et al., Semantic Uncertainty, ICLR 2023",
        "adaptation_note": (
            "표면 표현보다 의미 수준 비교가 중요하다는 관점 적용. "
            "SBERT cosine은 proxy 구현임 (원본 semantic entropy 재현 아님)."
        ),
    },
    "badge": {
        "official_repo": "https://github.com/JordanAsh/badge",
        "paper": "Ash et al., BADGE, ICLR 2020",
        "adaptation_note": "AL 방법론 lineage 명시. 본 스크립트는 gradient embedding 미구현.",
    },
    "coreset": {
        "official_repo": "https://github.com/ozansener/active_learning_coreset",
        "paper": "Sener & Savarese, Core-Set, ICLR 2018",
        "adaptation_note": "AL 방법론 lineage 명시. 본 스크립트는 Core-Set 최적화 미구현.",
    },
}


# =============================================================================
# [2] 데이터 클래스
# =============================================================================

@dataclass
class GenerationConfig:
    device: str
    model_id: str
    sbert_id: str
    seed: int
    do_sample: bool
    temperature: float
    top_p: float
    max_new_tokens: int


@dataclass
class ImageSample:
    """데이터셋에 관계없이 통일된 샘플 표현"""
    dataset: str          # "NEU_DET" | "KOLEKTOR" | "MVTEC"
    image_path: Path
    class_name: str       # NEU: defect class, Kolektor: "defect"/"ok", MVTec: category+type
    category: str         # MVTec category, 나머지는 class_name과 동일
    split: str            # "train" | "test"


# =============================================================================
# [3] 유틸 함수
# =============================================================================

def set_seed(seed: int) -> None:
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.backends.mps.is_available():
        torch.manual_seed(seed)


def prepare_output_dirs(dataset_tag: str) -> Dict[str, Path]:
    """
    출력 디렉토리 구조:
    OUTPUT_ROOT/
      EXPERIMENT_NAME/
        {dataset_tag}_{DATE_TAG}_{TIME_TAG}/
          jsonl/
          summary/
          plots/
          manifest/
          candidates/
          excel/
    """
    base_dir = OUTPUT_ROOT / EXPERIMENT_NAME / f"{dataset_tag}_{DATE_TAG}_{TIME_TAG}"
    dirs = {
        "base":       base_dir,
        "jsonl":      base_dir / "jsonl",
        "summary":    base_dir / "summary",
        "plots":      base_dir / "plots",
        "manifest":   base_dir / "manifest",
        "candidates": base_dir / "candidates",
        "excel":      base_dir / "excel",
    }
    for p in dirs.values():
        p.mkdir(parents=True, exist_ok=True)
    return dirs


# =============================================================================
# [4] 데이터셋별 이미지 샘플링
# =============================================================================

def sample_neu_det(root: Path, samples_per_class: int) -> List[ImageSample]:
    """
    NEU-DET: root/train/images/{class_dir}/*.jpg
    Source: http://faculty.neu.edu.cn/songkechen/zh_CN/zdylm/263270/list/index.htm
    6 classes: crazing, inclusion, patches, pitted_surface, rolled-in_scale, scratches
    """
    if not root.exists():
        raise FileNotFoundError(f"[NEU-DET] 경로 없음: {root}")

    # 클래스 폴더가 있는 경우 (계층 구조)
    class_dirs = sorted([d for d in root.iterdir() if d.is_dir()])
    samples = []

    if class_dirs:
        for class_dir in class_dirs:
            images = sorted([
                p for p in class_dir.iterdir()
                if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".bmp")
            ])
            if not images:
                continue
            chosen = images if len(images) <= samples_per_class else random.sample(images, samples_per_class)
            for img in chosen:
                samples.append(ImageSample(
                    dataset="NEU_DET",
                    image_path=img,
                    class_name=class_dir.name,
                    category=class_dir.name,
                    split="train",
                ))
    else:
        # 플랫 구조 (클래스 폴더 없이 파일명에 클래스 포함)
        images = sorted([
            p for p in root.iterdir()
            if p.suffix.lower() in (".jpg", ".jpeg", ".png", ".bmp")
        ])
        chosen = images if len(images) <= samples_per_class * 6 else random.sample(images, samples_per_class * 6)
        for img in chosen:
            # 파일명 앞부분에서 클래스 추출 (예: crazing_1.jpg → crazing)
            cls = "_".join(img.stem.split("_")[:-1]) if "_" in img.stem else img.stem
            samples.append(ImageSample(
                dataset="NEU_DET",
                image_path=img,
                class_name=cls,
                category=cls,
                split="train",
            ))
    return samples


def _parse_kolektor_annotation(ann_dir: Path, img_name: str) -> str:
    """
    KolektorSDD2-DatasetNinja 어노테이션 파싱.
    ann/{img_name}.json → objects 리스트 확인.
    Source: https://datasetninja.com/kolektor-surface-defect-dataset-2

    어노테이션 구조:
      { "objects": [ { "classTitle": "defect" | "ok", ... } ] }
    객체 없으면 "ok" (defect-free), 있으면 첫 번째 classTitle 반환.
    """
    ann_file = ann_dir / f"{img_name}.json"
    if not ann_file.exists():
        return "unknown"
    try:
        with open(ann_file, "r", encoding="utf-8") as f:
            ann = json.load(f)
        objects = ann.get("objects", [])
        if not objects:
            return "ok"
        return objects[0].get("classTitle", "defect")
    except Exception:
        return "unknown"


def sample_kolektor(root: Path, samples_per_class: int) -> List[ImageSample]:
    """
    KolektorSDD2-DatasetNinja:
      root/train/img/*.png  (이미지)
      root/train/ann/*.json (어노테이션)
    Source: https://www.vicos.si/resources/kolektorsdd2/
            Božič et al., Computers in Industry, 2021
    전략: defect / ok 클래스를 균형있게 샘플링.
    """
    if not root.exists():
        raise FileNotFoundError(f"[KOLEKTOR] 경로 없음: {root}")

    ann_dir = root.parent / "ann"

    images = sorted([
        p for p in root.iterdir()
        if p.suffix.lower() in (".png", ".jpg", ".jpeg", ".bmp")
    ])

    # 어노테이션 기반 클래스 분류
    by_class: Dict[str, List[Path]] = {}
    for img_path in images:
        cls = _parse_kolektor_annotation(ann_dir, img_path.name)
        by_class.setdefault(cls, []).append(img_path)

    samples = []
    for cls, img_list in sorted(by_class.items()):
        chosen = img_list if len(img_list) <= samples_per_class else random.sample(img_list, samples_per_class)
        for img in chosen:
            samples.append(ImageSample(
                dataset="KOLEKTOR",
                image_path=img,
                class_name=cls,
                category=cls,
                split="train",
            ))
    return samples


def sample_mvtec(root: Path, samples_per_class: int,
                 use_categories: Optional[List[str]],
                 use_test_split: bool) -> List[ImageSample]:
    """
    MVTec AD:
      root/{category}/train/good/*.png          (정상 이미지)
      root/{category}/test/{defect_type}/*.png  (결함 이미지)
    Source: https://www.mvtec.com/research-teaching/datasets/mvtec-ad
            Bergmann et al., CVPR 2019, CC BY-NC-SA 4.0

    전략:
    - use_test_split=True  → test/{defect_type}에서 결함 이미지 샘플링
    - use_test_split=False → train/good에서 정상 이미지 샘플링
    - use_categories가 지정되면 해당 카테고리만 사용
    """
    if not root.exists():
        raise FileNotFoundError(f"[MVTec] 경로 없음: {root}")

    all_categories = sorted([d.name for d in root.iterdir() if d.is_dir()])
    if use_categories:
        all_categories = [c for c in all_categories if c in use_categories]

    samples = []

    for category in all_categories:
        cat_root = root / category

        if use_test_split:
            test_root = cat_root / "test"
            if not test_root.exists():
                continue
            defect_dirs = sorted([
                d for d in test_root.iterdir()
                if d.is_dir() and d.name != "good"
            ])
            for defect_dir in defect_dirs:
                images = sorted([
                    p for p in defect_dir.iterdir()
                    if p.suffix.lower() in (".png", ".jpg", ".jpeg", ".bmp")
                ])
                if not images:
                    continue
                chosen = images if len(images) <= samples_per_class else random.sample(images, samples_per_class)
                for img in chosen:
                    samples.append(ImageSample(
                        dataset="MVTEC",
                        image_path=img,
                        class_name=f"{category}_{defect_dir.name}",
                        category=category,
                        split="test",
                    ))
        else:
            good_root = cat_root / "train" / "good"
            if not good_root.exists():
                continue
            images = sorted([
                p for p in good_root.iterdir()
                if p.suffix.lower() in (".png", ".jpg", ".jpeg", ".bmp")
            ])
            if not images:
                continue
            chosen = images if len(images) <= samples_per_class else random.sample(images, samples_per_class)
            for img in chosen:
                samples.append(ImageSample(
                    dataset="MVTEC",
                    image_path=img,
                    class_name=f"{category}_good",
                    category=category,
                    split="train",
                ))

    return samples


def sample_dataset(dataset_name: str) -> List[ImageSample]:
    """데이터셋 이름에 따라 적절한 샘플러 호출"""
    if dataset_name == "NEU_DET":
        return sample_neu_det(DATASET_PATHS["NEU_DET"], SAMPLES_PER_CLASS)
    elif dataset_name == "KOLEKTOR":
        return sample_kolektor(DATASET_PATHS["KOLEKTOR"], SAMPLES_PER_CLASS)
    elif dataset_name == "MVTEC":
        return sample_mvtec(
            DATASET_PATHS["MVTEC"],
            SAMPLES_PER_CLASS,
            MVTEC_USE_CATEGORIES,
            MVTEC_USE_TEST_SPLIT,
        )
    else:
        raise ValueError(f"알 수 없는 데이터셋: {dataset_name}")


# =============================================================================
# [5] 모델 로딩
# =============================================================================

def load_models():
    """
    Reference:
     - Qwen2-VL: https://huggingface.co/Qwen/Qwen2-VL-2B-Instruct
     - SBERT: https://sbert.net/
    """
    processor = AutoProcessor.from_pretrained(MODEL_ID, use_fast=False)
    vlm = Qwen2VLForConditionalGeneration.from_pretrained(
        MODEL_ID,
        torch_dtype=torch.float32,
        attn_implementation="eager",
    )
    vlm.to(DEVICE)
    vlm.eval()
    sbert = SentenceTransformer(SBERT_ID, device="cpu")
    return processor, vlm, sbert


# =============================================================================
# [6] VLM 응답 생성
# =============================================================================

def generate_response(
    image_path: Path,
    prompt: str,
    processor,
    vlm,
    gen_cfg: GenerationConfig,
) -> str:
    """
    Reference:
     - Qwen2-VL official example flow (apply_chat_template → process_vision_info → generate)
     - SelfCheckGPT: multi-generation consistency 아이디어 적용
    """
    messages = [
        {
            "role": "user",
            "content": [
                {"type": "image", "image": str(image_path)},
                {"type": "text",  "text":  prompt},
            ],
        }
    ]
    text = processor.apply_chat_template(messages, tokenize=False, add_generation_prompt=True)
    image_inputs, video_inputs = process_vision_info(messages)
    inputs = processor(
        text=[text],
        images=image_inputs,
        videos=video_inputs,
        padding=True,
        return_tensors="pt",
    )
    inputs = {k: v.to(DEVICE) for k, v in inputs.items()}

    with torch.no_grad():
        generated_ids = vlm.generate(
            **inputs,
            do_sample=gen_cfg.do_sample,
            temperature=gen_cfg.temperature,
            top_p=gen_cfg.top_p,
            max_new_tokens=gen_cfg.max_new_tokens,
        )

    generated_ids_trimmed = [
        out_ids[len(in_ids):]
        for in_ids, out_ids in zip(inputs["input_ids"], generated_ids)
    ]
    output_text = processor.batch_decode(
        generated_ids_trimmed,
        skip_special_tokens=True,
        clean_up_tokenization_spaces=True,
    )[0]
    return output_text.strip()


# =============================================================================
# [7] Consistency 계산
# =============================================================================

def cosine_similarity(a: np.ndarray, b: np.ndarray) -> float:
    a_norm = np.linalg.norm(a)
    b_norm = np.linalg.norm(b)
    if a_norm == 0.0 or b_norm == 0.0:
        return 0.0
    return float(np.dot(a, b) / (a_norm * b_norm))


def compute_sbert_consistency(
    responses: List[str], sbert
) -> Tuple[float, List[Dict]]:
    """
    Reference:
     - SBERT API: https://sbert.net/
     - SelfCheckGPT: multi-response consistency
     - Semantic Uncertainty: meaning-level comparison
    Note: SBERT cosine은 semantic entropy의 proxy 구현. 원본 재현 아님.
    """
    embeddings = sbert.encode(responses, convert_to_numpy=True, normalize_embeddings=True)
    pair_records = []
    scores = []
    for (i, emb_i), (j, emb_j) in itertools.combinations(enumerate(embeddings), 2):
        score = cosine_similarity(emb_i, emb_j)
        scores.append(score)
        pair_records.append({"pair": [i + 1, j + 1], "score": round(score, 6)})
    mean_score = float(np.mean(scores)) if scores else 0.0
    return mean_score, pair_records


def compute_bertscore_consistency(
    responses: List[str],
) -> Tuple[Optional[float], List[Dict]]:
    """
    Reference:
     - BERTScore paper: https://openreview.net/forum?id=SkeHuCVFDr
     - BERTScore repo: https://github.com/Tiiiger/bert_score
    Note: 계산 비용이 크므로 후처리 candidate에만 사용.
    """
    if bertscore_score is None:
        return None, []
    pair_records = []
    scores = []
    for (i, a), (j, b) in itertools.combinations(enumerate(responses), 2):
        _, _, f1 = bertscore_score([a], [b], lang="en", verbose=False, rescale_with_baseline=True)
        score = float(f1[0].item())
        scores.append(score)
        pair_records.append({"pair": [i + 1, j + 1], "score": round(score, 6)})
    mean_score = float(np.mean(scores)) if scores else None
    return mean_score, pair_records


# =============================================================================
# [8] 실험 실행
# =============================================================================

def run_prompt_group_experiment(
    samples: List[ImageSample],
    prompt_group_name: str,
    prompts: List[str],
    processor,
    vlm,
    sbert,
    gen_cfg: GenerationConfig,
    output_jsonl_path: Path,
) -> List[Dict]:
    """
    단일 prompt 그룹에 대해 전체 이미지 실험 수행.
    Reference:
     - SelfCheckGPT: https://github.com/potsawee/selfcheckgpt
     - Semantic Uncertainty: https://github.com/lorenzkuhn/semantic_uncertainty
    """
    records = []
    with open(output_jsonl_path, "w", encoding="utf-8") as f:
        for idx, sample in enumerate(samples, start=1):
            responses = []
            prompt_records = []
            for p_idx, prompt in enumerate(prompts, start=1):
                response = generate_response(sample.image_path, prompt, processor, vlm, gen_cfg)
                responses.append(response)
                prompt_records.append({
                    "prompt_index": p_idx,
                    "prompt_text":  prompt,
                    "response_text": response,
                })

            sbert_mean, sbert_pairs = compute_sbert_consistency(responses, sbert)
            if USE_BERTSCORE_FULL:
                bertscore_mean, bertscore_pairs = compute_bertscore_consistency(responses)
            else:
                bertscore_mean, bertscore_pairs = None, []

            record = {
                "experiment_name": EXPERIMENT_NAME,
                "timestamp":       datetime.now().isoformat(),
                "dataset":         sample.dataset,
                "prompt_group":    prompt_group_name,
                "image_index":     idx,
                "image_path":      str(sample.image_path),
                "class_name":      sample.class_name,
                "category":        sample.category,
                "split":           sample.split,
                "generation_config": asdict(gen_cfg),
                "method_references": {
                    "selfcheckgpt":        METHOD_UPSTREAM_SOURCES["selfcheckgpt"]["official_repo"],
                    "semantic_uncertainty": METHOD_UPSTREAM_SOURCES["semantic_uncertainty"]["official_repo"],
                },
                "prompt_records":          prompt_records,
                "sbert_pairwise":          sbert_pairs,
                "sbert_mean_score":        round(sbert_mean, 6),
                "bertscore_pairwise":      bertscore_pairs,
                "bertscore_mean_score":    round(bertscore_mean, 6) if bertscore_mean is not None else None,
            }
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
            records.append(record)

            print(
                f"[{sample.dataset}][{prompt_group_name}] "
                f"{idx}/{len(samples)} | {sample.class_name} | {sample.image_path.name} | "
                f"SBERT={sbert_mean:.4f}"
                + (f" | BERTScore={bertscore_mean:.4f}" if bertscore_mean is not None else "")
            )

    return records


# =============================================================================
# [9] 요약 통계
# =============================================================================

def summarize_records(records: List[Dict], metric_key: str) -> Dict:
    scores = [r[metric_key] for r in records if r.get(metric_key) is not None]
    summary = {
        "metric":  metric_key,
        "count":   len(scores),
        "mean":    round(float(np.mean(scores))   if scores else 0.0, 6),
        "std":     round(float(statistics.stdev(scores)) if len(scores) > 1 else 0.0, 6),
        "min":     round(float(np.min(scores))    if scores else 0.0, 6),
        "max":     round(float(np.max(scores))    if scores else 0.0, 6),
        "range":   round((float(np.max(scores)) - float(np.min(scores))) if scores else 0.0, 6),
        "class_stats": [],
    }
    by_class: Dict[str, List] = {}
    for r in records:
        if r.get(metric_key) is not None:
            by_class.setdefault(r["class_name"], []).append(r[metric_key])
    for cls, vals in sorted(by_class.items()):
        summary["class_stats"].append({
            "class_name": cls,
            "count":  len(vals),
            "mean":   round(float(np.mean(vals)), 6),
            "std":    round(float(statistics.stdev(vals)) if len(vals) > 1 else 0.0, 6),
            "min":    round(float(np.min(vals)), 6),
            "max":    round(float(np.max(vals)), 6),
        })
    return summary


# =============================================================================
# [10] 시각화
# =============================================================================

def plot_distributions(df: pd.DataFrame, metric_key: str, output_dir: Path) -> None:
    valid_df = df[df[metric_key].notna()].copy()
    if valid_df.empty:
        return

    # 히스토그램
    plt.figure(figsize=(8, 5))
    for group in sorted(valid_df["prompt_group"].unique()):
        subset = valid_df[valid_df["prompt_group"] == group][metric_key]
        plt.hist(subset, bins=15, alpha=0.5, label=group)
    plt.xlabel(metric_key)
    plt.ylabel("count")
    plt.title(f"{metric_key} histogram by prompt group")
    plt.legend()
    plt.tight_layout()
    plt.savefig(output_dir / f"{metric_key}_histogram.png", dpi=200)
    plt.close()

    # 클래스별 박스플롯
    groups  = sorted(valid_df["prompt_group"].unique())
    classes = sorted(valid_df["class_name"].unique())
    plt.figure(figsize=(max(12, len(classes) * 2), 6))
    positions, data, labels = [], [], []
    pos = 1
    for cls in classes:
        for group in groups:
            vals = valid_df[
                (valid_df["class_name"] == cls) &
                (valid_df["prompt_group"] == group)
            ][metric_key].values
            if len(vals) > 0:
                data.append(vals)
                positions.append(pos)
                labels.append(f"{cls}\n{group[:3]}")
                pos += 1
        pos += 1
    if data:
        plt.boxplot(data, positions=positions)
        plt.xticks(positions, labels, rotation=45, ha="right", fontsize=7)
        plt.ylabel(metric_key)
        plt.title(f"{metric_key} by class and prompt group")
        plt.tight_layout()
        plt.savefig(output_dir / f"{metric_key}_classwise_boxplot.png", dpi=200)
    plt.close()


# =============================================================================
# [11] Excel 저장 (test2604_revised.py 결과표와 동일 구조)
# =============================================================================

def save_excel(records: List[Dict], output_path: Path) -> None:
    """
    컬럼 구조:
     No | Dataset | Prompt Type | Class | Category | Image Name |
     SBERT | Prompt_1..5 | Response_1..5 | BERTScore

    test2604_revised.py 결과 엑셀과 동일한 형식 유지.
    날짜 prefix: output_path에 이미 포함됨.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # 헤더
    header_base = [
        "No", "Dataset", "Prompt_Type", "Class", "Category",
        "Image_Name", "SBERT_Mean",
    ]
    max_prompts = 5
    for i in range(1, max_prompts + 1):
        header_base.append(f"Prompt_{i}")
    for i in range(1, max_prompts + 1):
        header_base.append(f"Response_{i}")
    header_base.append("BERTScore_Mean")

    # 헤더 스타일
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(header_base, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # 데이터 행
    row_idx = 2
    # alternating 색상
    fill_even = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    fill_odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for rec_no, rec in enumerate(records, start=1):
        fill = fill_even if rec_no % 2 == 0 else fill_odd
        pr_map = {p["prompt_index"]: p for p in rec.get("prompt_records", [])}

        row_data = [
            rec_no,
            rec.get("dataset", ""),
            rec.get("prompt_group", ""),
            rec.get("class_name", ""),
            rec.get("category", ""),
            Path(rec.get("image_path", "")).name,
            rec.get("sbert_mean_score"),
        ]
        for i in range(1, max_prompts + 1):
            row_data.append(pr_map.get(i, {}).get("prompt_text", ""))
        for i in range(1, max_prompts + 1):
            row_data.append(pr_map.get(i, {}).get("response_text", ""))
        row_data.append(rec.get("bertscore_mean_score"))

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col_idx == 7:  # SBERT_Mean
                cell.number_format = "0.0000"

        row_idx += 1

    # 열 너비 자동 조정
    col_widths = {1: 6, 2: 12, 3: 12, 4: 18, 5: 14, 6: 20, 7: 12}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for i in range(8, 8 + max_prompts):
        ws.column_dimensions[get_column_letter(i)].width = 40
    for i in range(8 + max_prompts, 8 + max_prompts * 2):
        ws.column_dimensions[get_column_letter(i)].width = 50
    ws.column_dimensions[get_column_letter(8 + max_prompts * 2)].width = 14

    # 헤더 고정
    ws.freeze_panes = "A2"

    # 메타 시트
    ws_meta = wb.create_sheet("Meta")
    meta_rows = [
        ("Experiment",    EXPERIMENT_NAME),
        ("Date",          DATE_TAG),
        ("Time",          TIME_TAG),
        ("Model",         MODEL_ID),
        ("SBERT",         SBERT_ID),
        ("Seed",          SEED),
        ("Temperature",   TEMPERATURE),
        ("Top_p",         TOP_P),
        ("Max_new_tokens", MAX_NEW_TOKENS),
        ("SelfCheckGPT",  METHOD_UPSTREAM_SOURCES["selfcheckgpt"]["official_repo"]),
        ("SelfCheckGPT_paper", METHOD_UPSTREAM_SOURCES["selfcheckgpt"]["paper"]),
        ("SemanticUncertainty", METHOD_UPSTREAM_SOURCES["semantic_uncertainty"]["official_repo"]),
        ("SemanticUncertainty_paper", METHOD_UPSTREAM_SOURCES["semantic_uncertainty"]["paper"]),
        ("BADGE",         METHOD_UPSTREAM_SOURCES["badge"]["official_repo"]),
        ("CoreSet",       METHOD_UPSTREAM_SOURCES["coreset"]["official_repo"]),
        ("NEU-DET_source", "http://faculty.neu.edu.cn/songkechen/zh_CN/zdylm/263270/list/index.htm"),
        ("KolektorSDD2_source", "https://www.vicos.si/resources/kolektorsdd2/"),
        ("KolektorSDD2_paper", "Božič et al., Computers in Industry, 2021"),
        ("MVTec_source",  "https://www.mvtec.com/research-teaching/datasets/mvtec-ad"),
        ("MVTec_paper",   "Bergmann et al., CVPR 2019"),
        ("MVTec_license", "CC BY-NC-SA 4.0"),
    ]
    for r, (k, v) in enumerate(meta_rows, start=1):
        ws_meta.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws_meta.cell(row=r, column=2, value=v)
    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 70

    wb.save(output_path)
    print(f"  → Excel 저장: {output_path}")


# =============================================================================
# [12] Candidate Manifest (Active Learning용)
# =============================================================================

def save_candidate_manifests(
    records: List[Dict],
    metric_key: str,
    output_dir: Path,
    prefix: str,
) -> Dict:
    """
    Low/High/Random candidate set 저장.
    Reference:
     - BADGE: https://github.com/JordanAsh/badge
     - Core-Set: https://github.com/ozansener/active_learning_coreset
    """
    valid_records = [r for r in records if r.get(metric_key) is not None]
    sorted_records = sorted(valid_records, key=lambda x: x[metric_key])

    groups = {
        "low":    sorted_records[:LOW_K],
        "high":   sorted_records[-HIGH_K:],
        "random": random.sample(valid_records, min(RANDOM_K, len(valid_records))),
    }

    output_dir.mkdir(parents=True, exist_ok=True)
    result = {}
    for group_name, recs in groups.items():
        txt_path  = output_dir / f"{prefix}_{metric_key}_{group_name}_images.txt"
        json_path = output_dir / f"{prefix}_{metric_key}_{group_name}_records.json"
        with open(txt_path, "w", encoding="utf-8") as f:
            for rec in recs:
                f.write(f"{rec['image_path']}\n")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(recs, f, ensure_ascii=False, indent=2)
        result[group_name] = recs
    return result


def posthoc_bertscore_for_candidates(
    candidate_records: List[Dict],
    output_path: Path,
) -> List[Dict]:
    if not USE_BERTSCORE_CANDIDATES:
        return candidate_records
    if bertscore_score is None:
        print("BERTScore 패키지 미설치. 건너뜀.")
        return candidate_records

    updated = []
    for idx, rec in enumerate(candidate_records, start=1):
        responses = [p["response_text"] for p in rec["prompt_records"]]
        mean_score, pairwise = compute_bertscore_consistency(responses)
        new_rec = dict(rec)
        new_rec["posthoc_bertscore_mean_score"] = round(mean_score, 6) if mean_score else None
        new_rec["posthoc_bertscore_pairwise"]   = pairwise
        updated.append(new_rec)
        print(f"  [posthoc BERTScore] {idx}/{len(candidate_records)} → {new_rec['posthoc_bertscore_mean_score']}")

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(updated, f, ensure_ascii=False, indent=2)
    return updated


# =============================================================================
# [13] Manifest & JSON 저장
# =============================================================================

def save_source_manifest(
    output_path: Path,
    gen_cfg: GenerationConfig,
    samples: List[ImageSample],
    dataset_name: str,
) -> None:
    manifest = {
        "experiment_name": EXPERIMENT_NAME,
        "dataset":         dataset_name,
        "created_at":      datetime.now().isoformat(),
        "code_statement": (
            "이 스크립트는 upstream model card, API docs, method lineage 기반 실험 적용본입니다. "
            "원본 코드의 verbatim 복사가 아닙니다."
        ),
        "generation_config":    asdict(gen_cfg),
        "sampled_images_count": len(samples),
        "sampled_images":       [str(s.image_path) for s in samples],
        "method_upstream_sources": METHOD_UPSTREAM_SOURCES,
        "dataset_sources": {
            "NEU_DET":   "http://faculty.neu.edu.cn/songkechen/zh_CN/zdylm/263270/list/index.htm",
            "KOLEKTOR":  "https://www.vicos.si/resources/kolektorsdd2/ | Božič et al., CiI 2021",
            "MVTEC":     "https://www.mvtec.com/research-teaching/datasets/mvtec-ad | Bergmann et al., CVPR 2019 | CC BY-NC-SA 4.0",
        },
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)


def save_json(data: Dict, output_path: Path) -> None:
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def records_to_dataframe(records: List[Dict]) -> pd.DataFrame:
    rows = []
    for rec in records:
        row = {
            "experiment_name":   rec["experiment_name"],
            "timestamp":         rec["timestamp"],
            "dataset":           rec.get("dataset", ""),
            "prompt_group":      rec["prompt_group"],
            "image_index":       rec["image_index"],
            "image_path":        rec["image_path"],
            "class_name":        rec["class_name"],
            "category":          rec.get("category", ""),
            "split":             rec.get("split", ""),
            "sbert_mean_score":  rec["sbert_mean_score"],
            "bertscore_mean_score": rec.get("bertscore_mean_score"),
        }
        for pr in rec["prompt_records"]:
            row[f"prompt_{pr['prompt_index']}"]   = pr["prompt_text"]
            row[f"response_{pr['prompt_index']}"] = pr["response_text"]
        rows.append(row)
    return pd.DataFrame(rows)


# =============================================================================
# [14] 데이터셋별 실험 실행
# =============================================================================

def run_dataset_experiment(
    dataset_name: str,
    processor,
    vlm,
    sbert,
    gen_cfg: GenerationConfig,
) -> None:
    print(f"\n{'='*60}")
    print(f"[실험 시작] Dataset: {dataset_name}  ({DATE_TAG})")
    print(f"{'='*60}")

    dirs = prepare_output_dirs(dataset_name)

    # 이미지 샘플링
    try:
        samples = sample_dataset(dataset_name)
    except FileNotFoundError as e:
        print(f"  ⚠ 데이터 경로 없음, 건너뜀: {e}")
        return

    print(f"  샘플 수: {len(samples)}개")
    save_source_manifest(
        dirs["manifest"] / f"source_manifest_{DATE_TAG}.json",
        gen_cfg, samples, dataset_name,
    )

    all_records = []

    # ── Rephrase 실험 ─────────────────────────────────────────────────────
    rephrase_records = run_prompt_group_experiment(
        samples=samples,
        prompt_group_name="rephrase",
        prompts=REPHRASE_PROMPTS,
        processor=processor,
        vlm=vlm,
        sbert=sbert,
        gen_cfg=gen_cfg,
        output_jsonl_path=dirs["jsonl"] / f"rephrase_results_{DATE_TAG}.jsonl",
    )
    all_records.extend(rephrase_records)

    # ── Viewpoint 실험 ────────────────────────────────────────────────────
    viewpoint_records = run_prompt_group_experiment(
        samples=samples,
        prompt_group_name="viewpoint",
        prompts=VIEWPOINT_PROMPTS,
        processor=processor,
        vlm=vlm,
        sbert=sbert,
        gen_cfg=gen_cfg,
        output_jsonl_path=dirs["jsonl"] / f"viewpoint_results_{DATE_TAG}.jsonl",
    )
    all_records.extend(viewpoint_records)

    # ── DataFrame & CSV ───────────────────────────────────────────────────
    df = records_to_dataframe(all_records)
    csv_path = dirs["summary"] / f"results_{dataset_name}_{DATE_TAG}.csv"
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"  → CSV 저장: {csv_path}")

    # ── Excel 저장 ────────────────────────────────────────────────────────
    excel_path = dirs["excel"] / f"results_{dataset_name}_{DATE_TAG}.xlsx"
    save_excel(all_records, excel_path)

    # ── 요약 통계 ──────────────────────────────────────────────────────────
    sbert_summary = summarize_records(all_records, "sbert_mean_score")
    save_json(sbert_summary, dirs["summary"] / f"sbert_summary_{dataset_name}_{DATE_TAG}.json")
    print(f"  SBERT 전체 평균: {sbert_summary['mean']:.4f} (n={sbert_summary['count']})")

    # ── 시각화 ────────────────────────────────────────────────────────────
    plot_distributions(df, "sbert_mean_score", dirs["plots"])

    # ── Candidate Manifest (Active Learning용) ───────────────────────────
    for group_name, group_records in [
        ("rephrase",  rephrase_records),
        ("viewpoint", viewpoint_records),
    ]:
        candidates = save_candidate_manifests(
            records=group_records,
            metric_key="sbert_mean_score",
            output_dir=dirs["candidates"],
            prefix=f"{dataset_name}_{group_name}_{DATE_TAG}",
        )
        # 후처리 BERTScore (low + high에만)
        for tier in ("low", "high"):
            if candidates[tier]:
                posthoc_path = (
                    dirs["candidates"]
                    / f"{dataset_name}_{group_name}_{DATE_TAG}_bertscore_{tier}.json"
                )
                posthoc_bertscore_for_candidates(candidates[tier], posthoc_path)

    print(f"\n[완료] {dataset_name} 결과 저장 위치: {dirs['base']}")


# =============================================================================
# [15] main
# =============================================================================

def main():
    print(f"\n{'#'*60}")
    print(f"# VLM Defect Consistency 실험  |  {DATE_TAG}")
    print(f"# 실험명: {EXPERIMENT_NAME}")
    print(f"# 대상 데이터셋: {ACTIVE_DATASETS}")
    print(f"{'#'*60}\n")

    set_seed(SEED)

    gen_cfg = GenerationConfig(
        device=DEVICE,
        model_id=MODEL_ID,
        sbert_id=SBERT_ID,
        seed=SEED,
        do_sample=DO_SAMPLE,
        temperature=TEMPERATURE,
        top_p=TOP_P,
        max_new_tokens=MAX_NEW_TOKENS,
    )

    print("모델 로딩 중...")
    processor, vlm, sbert = load_models()
    print("모델 로딩 완료.\n")

    for dataset_name in ACTIVE_DATASETS:
        run_dataset_experiment(dataset_name, processor, vlm, sbert, gen_cfg)

    print(f"\n{'='*60}")
    print(f"전체 실험 완료. 날짜: {DATE_TAG}")
    print(f"결과 루트: {OUTPUT_ROOT / EXPERIMENT_NAME}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()