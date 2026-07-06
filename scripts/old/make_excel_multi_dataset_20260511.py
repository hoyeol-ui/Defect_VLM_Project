import csv
import glob
from pathlib import Path
from statistics import mean, stdev
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule
from PIL import Image as PILImage


# =============================================================================
# [1] 경로 설정 — 실제 환경에 맞게 수정하세요
# =============================================================================

LOGS_ROOT = Path(
    "/Users/hy/PycharmProjects/PythonProject"
    "/Defect_VLM_Project/logs/consistency_multi_dataset_20260511"
)

DATE_TAG = "20260511"

# CSV 파일 자동 탐색 (timestamp 폴더명이 달라도 자동으로 찾음)
def _find_csv(dataset_name: str) -> Optional[Path]:
    """
    LOGS_ROOT/{dataset_name}_*/summary/results_{dataset_name}_{DATE_TAG}.csv
    패턴으로 자동 탐색. 여러 개면 가장 최신(알파벳 마지막) 사용.
    """
    pattern = str(LOGS_ROOT / f"{dataset_name}_{DATE_TAG}_*" / "summary" / f"results_{dataset_name}_{DATE_TAG}.csv")
    matches = sorted(glob.glob(pattern))
    if matches:
        return Path(matches[-1])
    # fallback: 직접 지정 방식
    return None

# 직접 경로 지정 (자동 탐색 실패 시 여기에 절대 경로 입력)
CSV_PATHS: Dict[str, Optional[Path]] = {
    "NEU_DET":  _find_csv("NEU_DET"),
    "KOLEKTOR": _find_csv("KOLEKTOR"),
    "MVTEC":    _find_csv("MVTEC"),
}

OUTPUT_XLSX = LOGS_ROOT / f"combined_review_{DATE_TAG}.xlsx"
THUMB_DIR   = LOGS_ROOT / "_thumbs_combined"
THUMB_DIR.mkdir(exist_ok=True)

# =============================================================================
# [2] 스타일 상수
# =============================================================================

THIN   = Side(style="thin",   color="D9D9D9")
MEDIUM = Side(style="medium", color="A6A6A6")
BORDER_THIN   = Border(top=THIN,   bottom=THIN,   left=THIN,   right=THIN)
BORDER_MEDIUM = Border(top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=MEDIUM)

HEADER_FILL     = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT     = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL  = PatternFill("solid", fgColor="D9EAF7")
META_FILL       = PatternFill("solid", fgColor="F7F7F7")
PROMPT_FILL     = PatternFill("solid", fgColor="FFF2CC")
RESPONSE_FILL   = PatternFill("solid", fgColor="F2F2F2")
NOTE_FILL       = PatternFill("solid", fgColor="FFF9E6")

# 데이터셋별 색상
DATASET_FILL = {
    "NEU_DET":  "DDEBF7",   # 파랑 계열
    "KOLEKTOR": "FCE4D6",   # 주황 계열
    "MVTEC":    "E2EFDA",   # 초록 계열
}

# Prompt type 색상
PROMPT_TYPE_FILL = {
    "rephrase":  "BDD7EE",
    "viewpoint": "F4B183",
}


# =============================================================================
# [3] CSV 로딩 → 통일된 row dict 변환
# =============================================================================

def load_csv_rows(csv_path: Path, dataset_name: str) -> List[Dict]:
    """
    CSV 컬럼:
      experiment_name, timestamp, dataset, prompt_group, image_index,
      image_path, class_name, category, split, sbert_mean_score,
      bertscore_mean_score,
      prompt_1..5, response_1..5
    """
    rows = []
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for idx, rec in enumerate(reader, start=1):
            image_path = Path(rec.get("image_path", ""))

            # prompt_records 구성
            prompt_records = []
            for i in range(1, 6):
                p = rec.get(f"prompt_{i}", "")
                r = rec.get(f"response_{i}", "")
                if p or r:
                    prompt_records.append({
                        "prompt_index": i,
                        "prompt_text":  p,
                        "response_text": r,
                    })

            sbert_val = rec.get("sbert_mean_score", "")
            try:
                sbert_float = float(sbert_val) if sbert_val else None
            except ValueError:
                sbert_float = None

            rows.append({
                "global_no":    idx,          # 임시, 나중에 재부여
                "dataset":      dataset_name,
                "prompt_group": rec.get("prompt_group", ""),
                "image_index":  rec.get("image_index", ""),
                "image_path":   str(image_path),
                "image_name":   image_path.name,
                "class_name":   rec.get("class_name", ""),
                "category":     rec.get("category", ""),
                "split":        rec.get("split", ""),
                "sbert_mean":   sbert_float,
                "pair_min":     None,   # CSV에 없음
                "pair_max":     None,
                "prompt_records": prompt_records,
            })
    return rows


def load_all_rows() -> List[Dict]:
    all_rows = []
    for dataset_name, csv_path in CSV_PATHS.items():
        if csv_path is None or not csv_path.exists():
            print(f"  ⚠ [{dataset_name}] CSV 없음, 건너뜀: {csv_path}")
            continue
        rows = load_csv_rows(csv_path, dataset_name)
        print(f"  [{dataset_name}] {len(rows)}행 로드: {csv_path.name}")
        all_rows.extend(rows)

    # global_no 재부여
    for i, row in enumerate(all_rows, start=1):
        row["global_no"] = i

    return all_rows


# =============================================================================
# [4] 썸네일 생성
# =============================================================================

def make_thumbnail(image_path_str: str, size=(160, 160)) -> Optional[Path]:
    image_path = Path(image_path_str)
    if not image_path.exists():
        return None
    thumb_path = THUMB_DIR / f"{image_path.stem}_{image_path.parent.name}{image_path.suffix}"
    try:
        img = PILImage.open(image_path).convert("RGB")
        img.thumbnail(size)
        img.save(thumb_path)
        return thumb_path
    except Exception as e:
        print(f"  썸네일 오류: {image_path.name} | {e}")
        return None


# =============================================================================
# [5] 스타일 유틸
# =============================================================================

def style_cell(cell, *, fill=None, font=None, alignment=None, border=None, number_format=None):
    if fill is not None:        cell.fill = fill
    if font is not None:        cell.font = font
    if alignment is not None:   cell.alignment = alignment
    if border is not None:      cell.border = border
    if number_format is not None: cell.number_format = number_format


def fill_merged_range(ws, cell_range, fill=None, border=None, font=None, alignment=None):
    for row in ws[cell_range]:
        for cell in row:
            style_cell(cell, fill=fill, border=border, font=font, alignment=alignment)


# =============================================================================
# [6] Sheet 1 — Dashboard
# =============================================================================

def build_dashboard(wb, rows: List[Dict]):
    ws = wb.active
    ws.title = "dashboard"

    # 타이틀
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Multi-Dataset Consistency Review Dashboard  |  {DATE_TAG}"
    style_cell(ws["A1"],
               font=Font(size=15, bold=True, color="FFFFFF"),
               fill=HEADER_FILL,
               alignment=Alignment(horizontal="center", vertical="center"))
    fill_merged_range(ws, "A1:J1", fill=HEADER_FILL, border=BORDER_MEDIUM)
    ws.row_dimensions[1].height = 28

    # 범례
    ws["A3"] = "Dataset Color Legend"
    style_cell(ws["A3"], font=Font(bold=True), fill=SUBHEADER_FILL, border=BORDER_THIN)
    ws.merge_cells("A3:J3")
    fill_merged_range(ws, "A3:J3", fill=SUBHEADER_FILL, border=BORDER_THIN)

    legend_items = [
        ("NEU_DET",  DATASET_FILL["NEU_DET"],  "Steel surface, 6 classes (crazing/inclusion/patches/pitted_surface/rolled-in_scale/scratches)"),
        ("KOLEKTOR", DATASET_FILL["KOLEKTOR"], "Production surface, binary (defect / ok)"),
        ("MVTEC",    DATASET_FILL["MVTEC"],    "15 categories, anomaly detection benchmark"),
    ]
    for i, (name, color, desc) in enumerate(legend_items, start=4):
        ws.cell(i, 1, name)
        ws.cell(i, 2, desc)
        style_cell(ws.cell(i, 1), fill=PatternFill("solid", fgColor=color), font=Font(bold=True), border=BORDER_THIN,
                   alignment=Alignment(horizontal="center", vertical="center"))
        style_cell(ws.cell(i, 2), border=BORDER_THIN, alignment=Alignment(wrap_text=True))
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=10)
        fill_merged_range(ws, f"B{i}:J{i}", border=BORDER_THIN)

    # 요약 테이블
    header_row = 9
    headers = [
        "Dataset", "Prompt Group", "Class", "Count",
        "SBERT Mean", "SBERT Std", "SBERT Min", "SBERT Max", "SBERT Range",
        "Source"
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(header_row, col_idx, h)
        style_cell(cell, fill=HEADER_FILL, font=HEADER_FONT,
                   alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                   border=BORDER_MEDIUM)
    ws.row_dimensions[header_row].height = 24

    dataset_sources = {
        "NEU_DET":  "http://faculty.neu.edu.cn/songkechen/zh_CN/zdylm/263270/list/index.htm",
        "KOLEKTOR": "https://www.vicos.si/resources/kolektorsdd2/",
        "MVTEC":    "https://www.mvtec.com/research-teaching/datasets/mvtec-ad",
    }

    cur = header_row + 1
    # 데이터셋 × prompt_group 기준으로 집계
    from itertools import groupby
    sorted_rows = sorted(rows, key=lambda r: (r["dataset"], r["prompt_group"], r["class_name"]))

    written_ds = set()
    for (ds, pg), group_iter in groupby(sorted_rows, key=lambda r: (r["dataset"], r["prompt_group"])):
        group_list = list(group_iter)
        # class별 sub-grouping
        by_class = {}
        for r in group_list:
            by_class.setdefault(r["class_name"], []).append(r)

        ds_fill = PatternFill("solid", fgColor=DATASET_FILL.get(ds, "FFFFFF"))
        pg_fill = PatternFill("solid", fgColor=PROMPT_TYPE_FILL.get(pg, "FFFFFF"))

        for cls, cls_rows in sorted(by_class.items()):
            sbert_vals = [r["sbert_mean"] for r in cls_rows if r["sbert_mean"] is not None]
            row_data = [
                ds, pg, cls, len(cls_rows),
                round(mean(sbert_vals), 4) if sbert_vals else None,
                round(stdev(sbert_vals), 4) if len(sbert_vals) > 1 else None,
                round(min(sbert_vals), 4) if sbert_vals else None,
                round(max(sbert_vals), 4) if sbert_vals else None,
                round(max(sbert_vals) - min(sbert_vals), 4) if sbert_vals else None,
                dataset_sources.get(ds, ""),
            ]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws.cell(cur, col_idx, val)
                if col_idx == 1:
                    style_cell(cell, fill=ds_fill, border=BORDER_THIN,
                               alignment=Alignment(horizontal="center", vertical="center"))
                elif col_idx == 2:
                    style_cell(cell, fill=pg_fill, border=BORDER_THIN,
                               alignment=Alignment(horizontal="center", vertical="center"))
                else:
                    style_cell(cell, border=BORDER_THIN,
                               alignment=Alignment(horizontal="center", vertical="center", wrap_text=True))
                if col_idx in (5, 6, 7, 8, 9):
                    cell.number_format = "0.0000"
            cur += 1

        # 데이터셋 소계
        ds_rows = [r for r in rows if r["dataset"] == ds and r["prompt_group"] == pg]
        sbert_all = [r["sbert_mean"] for r in ds_rows if r["sbert_mean"] is not None]
        ws.cell(cur, 1, f"[{ds} / {pg} 전체]")
        ws.cell(cur, 4, len(ds_rows))
        if sbert_all:
            ws.cell(cur, 5, round(mean(sbert_all), 4))
            ws.cell(cur, 6, round(stdev(sbert_all), 4) if len(sbert_all) > 1 else None)
            ws.cell(cur, 7, round(min(sbert_all), 4))
            ws.cell(cur, 8, round(max(sbert_all), 4))
            ws.cell(cur, 9, round(max(sbert_all) - min(sbert_all), 4))
        subtotal_fill = PatternFill("solid", fgColor="E2E2E2")
        for c in range(1, 11):
            style_cell(ws.cell(cur, c),
                       fill=subtotal_fill,
                       font=Font(bold=True),
                       border=BORDER_THIN,
                       alignment=Alignment(horizontal="center", vertical="center"))
            if c in (5, 6, 7, 8, 9):
                ws.cell(cur, c).number_format = "0.0000"
        cur += 1

    # 컬럼 너비
    col_widths = {1: 14, 2: 14, 3: 24, 4: 8, 5: 12, 6: 12, 7: 12, 8: 12, 9: 12, 10: 52}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # SBERT 컬러스케일
    if cur > header_row + 2:
        ws.conditional_formatting.add(
            f"E{header_row+1}:H{cur-1}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            )
        )
    ws.freeze_panes = "A9"


# =============================================================================
# [7] Sheet 2 — Review Cards (이미지 + prompt/response 카드)
# =============================================================================

def build_review_cards(wb, rows: List[Dict]):
    ws = wb.create_sheet("review_cards")

    # 컬럼 너비 — 기존 make_excel_visual_v2 구조 유지
    col_widths = {
        "A": 6,   # No
        "B": 22,  # Thumbnail
        "C": 14,  # Dataset
        "D": 13,  # Prompt Type
        "E": 10,  # Class
        "F": 10,  # Category
        "G": 22,  # Image Name
        "H": 10,  # SBERT
        "I": 10,  # Pair Min
        "J": 10,  # Pair Max
        "K": 6,   # P#
        "L": 36,  # Prompt
        "M": 52,  # Response
        "N": 11,  # Semantic
        "O": 11,  # Spatial
        "P": 11,  # Attribute
        "Q": 11,  # Domain
        "R": 10,  # Total
        "S": 14,  # Final Use
        "T": 32,  # Manual Note
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    headers = [
        "No", "Thumbnail", "Dataset", "Prompt Type", "Class", "Category",
        "Image Name", "SBERT", "Pair Min", "Pair Max", "P#",
        "Prompt", "Response",
        "Semantic", "Spatial", "Attribute", "Domain", "Total", "Final Use", "Manual Note"
    ]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(1, i, h)
        style_cell(cell,
                   fill=HEADER_FILL, font=HEADER_FONT,
                   alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                   border=BORDER_MEDIUM)
    ws.row_dimensions[1].height = 28

    # 드롭다운
    score_dv = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True)
    final_dv = DataValidation(type="list", formula1='"YES,NO,MAYBE"', allow_blank=True)
    ws.add_data_validation(score_dv)
    ws.add_data_validation(final_dv)

    row_ptr = 2

    # 데이터셋별로 섹션 구분
    from itertools import groupby
    sorted_rows = sorted(rows, key=lambda r: (r["dataset"], r["prompt_group"], r["global_no"]))

    for (ds, pg), group_iter in groupby(sorted_rows, key=lambda r: (r["dataset"], r["prompt_group"])):
        group_list = list(group_iter)
        ds_fill   = PatternFill("solid", fgColor=DATASET_FILL.get(ds, "DDDDDD"))
        pg_fill   = PatternFill("solid", fgColor=PROMPT_TYPE_FILL.get(pg, "DDDDDD"))

        # 섹션 헤더
        ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=20)
        ws.cell(row_ptr, 1,
                f"  ■  {ds}  |  prompt: {pg}  |  {len(group_list)} samples")
        style_cell(ws.cell(row_ptr, 1),
                   fill=ds_fill,
                   font=Font(bold=True, size=11),
                   alignment=Alignment(horizontal="left", vertical="center"),
                   border=BORDER_MEDIUM)
        fill_merged_range(ws, f"A{row_ptr}:T{row_ptr}", fill=ds_fill, border=BORDER_MEDIUM)
        ws.row_dimensions[row_ptr].height = 22
        row_ptr += 1

        # 각 이미지 카드
        for row in group_list:
            prompts = row["prompt_records"]
            num_rows = max(5, len(prompts))
            block_start = row_ptr
            block_end   = row_ptr + num_rows - 1

            # 병합 컬럼: No, Thumbnail, Dataset, Prompt Type, Class, Category,
            #            Image Name, SBERT, Pair Min, Pair Max, Total, Final Use, Note
            merge_cols = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 18, 19, 20]
            for col in merge_cols:
                if block_end > block_start:
                    ws.merge_cells(start_row=block_start, start_column=col,
                                   end_row=block_end,   end_column=col)

            # 값 입력
            ws.cell(block_start, 1,  row["global_no"])
            ws.cell(block_start, 3,  row["dataset"])
            ws.cell(block_start, 4,  row["prompt_group"])
            ws.cell(block_start, 5,  row["class_name"])
            ws.cell(block_start, 6,  row["category"])
            ws.cell(block_start, 7,  row["image_name"])
            ws.cell(block_start, 8,  row["sbert_mean"])
            ws.cell(block_start, 9,  row["pair_min"])
            ws.cell(block_start, 10, row["pair_max"])
            ws.cell(block_start, 18, f"=SUM(N{block_start}:Q{block_start})")
            ws.cell(block_start, 19, "MAYBE")
            ws.cell(block_start, 20, "")

            # 셀 스타일 — 메타 열
            for col in [1, 5, 6, 7, 8, 9, 10, 18, 19, 20]:
                style_cell(ws.cell(block_start, col),
                           fill=META_FILL,
                           alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                           border=BORDER_THIN)

            # Dataset / Prompt Type 열 색상
            fill_merged_range(ws, f"C{block_start}:C{block_end}", fill=ds_fill,  border=BORDER_THIN)
            fill_merged_range(ws, f"D{block_start}:D{block_end}", fill=pg_fill,  border=BORDER_THIN)

            # 셀 정렬
            for col in [3, 4]:
                ws.cell(block_start, col).alignment = Alignment(horizontal="center", vertical="center")

            # SBERT 노란 배경
            for col_ref in [f"H{block_start}", f"I{block_start}", f"J{block_start}"]:
                ws[col_ref].fill = NOTE_FILL

            # 숫자 포맷
            for col in [8, 9, 10]:
                ws.cell(block_start, col).number_format = "0.0000"

            # 하이퍼링크
            ws.cell(block_start, 7).hyperlink = row["image_path"]
            ws.cell(block_start, 7).style = "Hyperlink"

            # 드롭다운 범위 등록
            score_dv.add(f"N{block_start}:Q{block_start}")
            final_dv.add(f"S{block_start}")

            # Prompt / Response 행
            for i in range(num_rows):
                cur_r = block_start + i
                ws.row_dimensions[cur_r].height = 36

                # P# 번호
                ws.cell(cur_r, 11, i + 1)
                style_cell(ws.cell(cur_r, 11),
                           fill=SUBHEADER_FILL,
                           alignment=Alignment(horizontal="center", vertical="center"),
                           border=BORDER_THIN)

                pt = prompts[i]["prompt_text"]  if i < len(prompts) else ""
                rt = prompts[i]["response_text"] if i < len(prompts) else ""
                ws.cell(cur_r, 12, pt)
                ws.cell(cur_r, 13, rt)
                style_cell(ws.cell(cur_r, 12), fill=PROMPT_FILL,
                           alignment=Alignment(vertical="center", wrap_text=True), border=BORDER_THIN)
                style_cell(ws.cell(cur_r, 13), fill=RESPONSE_FILL,
                           alignment=Alignment(vertical="center", wrap_text=True), border=BORDER_THIN)

            # 썸네일 삽입
            thumb = make_thumbnail(row["image_path"])
            if thumb:
                try:
                    xl_img = XLImage(str(thumb))
                    xl_img.anchor = f"B{block_start}"
                    ws.add_image(xl_img)
                    ws.row_dimensions[block_start].height = max(
                        ws.row_dimensions[block_start].height, 130
                    )
                except Exception as e:
                    print(f"  이미지 삽입 오류: {row['image_name']} | {e}")

            # 전체 border
            for r in range(block_start, block_end + 1):
                for c in range(1, 21):
                    ws.cell(r, c).border = BORDER_THIN
            # 블록 상단/하단 굵은 선
            for c in range(1, 21):
                ws.cell(block_start, c).border = Border(
                    top=MEDIUM, bottom=THIN, left=THIN, right=THIN)
                ws.cell(block_end,   c).border = Border(
                    top=THIN, bottom=MEDIUM, left=THIN, right=THIN)

            row_ptr = block_end + 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:T{max(2, row_ptr - 1)}"
    if row_ptr > 3:
        ws.conditional_formatting.add(
            f"H2:H{row_ptr-1}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            )
        )


# =============================================================================
# [8] Sheet 3 — Review Table (플랫 테이블)
# =============================================================================

def build_review_table(wb, rows: List[Dict]):
    ws = wb.create_sheet("review_table")

    headers = [
        "No", "Dataset", "Prompt Group", "Class", "Category", "Split",
        "Image Name", "Image Path", "SBERT Mean",
        "Prompt 1", "Response 1",
        "Prompt 2", "Response 2",
        "Prompt 3", "Response 3",
        "Prompt 4", "Response 4",
        "Prompt 5", "Response 5",
        "Semantic", "Spatial", "Attribute", "Domain", "Total", "Final Use", "Manual Note"
    ]
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(1, idx, h)
        style_cell(cell,
                   fill=HEADER_FILL, font=HEADER_FONT,
                   alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                   border=BORDER_MEDIUM)
    ws.row_dimensions[1].height = 26

    score_dv = DataValidation(type="list", formula1='"0,1,2"', allow_blank=True)
    final_dv = DataValidation(type="list", formula1='"YES,NO,MAYBE"', allow_blank=True)
    ws.add_data_validation(score_dv)
    ws.add_data_validation(final_dv)

    for row_idx, row in enumerate(rows, start=2):
        prompts = row["prompt_records"]
        values = [
            row["global_no"],
            row["dataset"],
            row["prompt_group"],
            row["class_name"],
            row["category"],
            row["split"],
            row["image_name"],
            row["image_path"],
            row["sbert_mean"],
        ]
        for i in range(5):
            values.append(prompts[i]["prompt_text"]   if i < len(prompts) else "")
            values.append(prompts[i]["response_text"] if i < len(prompts) else "")
        # Semantic, Spatial, Attribute, Domain, Total, Final Use, Note
        values.extend([None, None, None, None,
                       f"=SUM(T{row_idx}:W{row_idx})",
                       "MAYBE", ""])

        for col_idx, val in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, val)
            ws.cell(row_idx, col_idx).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row_idx, col_idx).border = BORDER_THIN

        # 데이터셋 색상
        ds_fill = PatternFill("solid", fgColor=DATASET_FILL.get(row["dataset"], "FFFFFF"))
        pg_fill = PatternFill("solid", fgColor=PROMPT_TYPE_FILL.get(row["prompt_group"], "FFFFFF"))
        ws.cell(row_idx, 2).fill = ds_fill
        ws.cell(row_idx, 3).fill = pg_fill

        # 숫자 포맷
        ws.cell(row_idx, 9).number_format = "0.0000"

        # 하이퍼링크
        ws.cell(row_idx, 7).hyperlink = row["image_path"]
        ws.cell(row_idx, 7).style = "Hyperlink"

        # 드롭다운
        score_dv.add(f"T{row_idx}:W{row_idx}")
        final_dv.add(f"Y{row_idx}")

    widths = {
        1: 6,  2: 12,  3: 13,  4: 20,  5: 12,  6: 8,
        7: 22,  8: 52,  9: 12,
        10: 30, 11: 44, 12: 30, 13: 44, 14: 30, 15: 44,
        16: 30, 17: 44, 18: 30, 19: 44,
        20: 10, 21: 10, 22: 10, 23: 10, 24: 10, 25: 12, 26: 28,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    max_row = max(2, len(rows) + 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max_row}"

    if len(rows) > 0:
        ws.conditional_formatting.add(
            f"I2:I{max_row}",
            ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B",
            )
        )
        ws.conditional_formatting.add(
            f"X2:X{max_row}",
            ColorScaleRule(
                start_type="min", start_color="63BE7B",
                mid_type="num", mid_value=4, mid_color="FFEB84",
                end_type="max", end_color="F8696B",
            )
        )


# =============================================================================
# [9] Sheet 4 — Scoring Guide
# =============================================================================

def build_scoring_guide(wb):
    ws = wb.create_sheet("scoring_guide")

    guide_rows = [
        ["Criterion", "0", "1", "2", "Interpretation"],
        ["Semantic Inconsistency",
         "Same defect meaning",
         "Minor semantic variation",
         "Entirely different defect types",
         "예: dark spot ↔ stain → 1,  dark spot ↔ crack → 2"],
        ["Spatial/Direction Inconsistency",
         "Same location/direction",
         "Minor positional difference",
         "Contradictory location or direction",
         "예: vertical ↔ horizontal → 2"],
        ["Attribute Inconsistency",
         "Same shape/texture",
         "Minor attribute difference",
         "Different shape/texture",
         "예: circular spot ↔ long streak → 2"],
        ["Domain Error",
         "No domain error",
         "Weak or questionable domain term",
         "Clear wrong domain",
         "예: metal defect인데 fabric/textile 표현이 나오면 2"],
        ["Total Difficulty",
         "0–2: easy",
         "3–5: moderate",
         "6–8: difficult",
         "AL 후보 우선순위 판단 기준"],
        ["Final Use for AL?",
         "YES: 레이블링 필요",
         "MAYBE: 검토 후 결정",
         "NO: 불필요",
         "라벨링 예산 배분 기준"],
        [],
        ["Dataset Source", "", "", "", ""],
        ["NEU-DET",
         "http://faculty.neu.edu.cn/songkechen/zh_CN/zdylm/263270/list/index.htm",
         "Steel surface, 6 classes", "1,800 images", ""],
        ["KolektorSDD2",
         "https://www.vicos.si/resources/kolektorsdd2/",
         "Božič et al., Computers in Industry 2021",
         "3,335 images, binary (defect/ok)", "CC BY-NC-SA 4.0"],
        ["MVTec AD",
         "https://www.mvtec.com/research-teaching/datasets/mvtec-ad",
         "Bergmann et al., CVPR 2019",
         "15 categories, 5,354 images", "CC BY-NC-SA 4.0"],
    ]

    for row_data in guide_rows:
        ws.append(row_data)

    # 헤더
    for cell in ws[1]:
        style_cell(cell,
                   fill=HEADER_FILL, font=HEADER_FONT,
                   alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                   border=BORDER_MEDIUM)

    # 내용 행
    for r in range(2, len(guide_rows) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, 6)]
        if not any(row_vals):
            continue
        if r == 9:  # Dataset Source 소제목
            for c in range(1, 6):
                style_cell(ws.cell(r, c),
                           fill=SUBHEADER_FILL, font=Font(bold=True),
                           border=BORDER_THIN,
                           alignment=Alignment(horizontal="center", vertical="center"))
            continue
        for c in range(1, 6):
            style_cell(ws.cell(r, c),
                       border=BORDER_THIN,
                       alignment=Alignment(vertical="center", wrap_text=True))
        ws.cell(r, 1).fill = SUBHEADER_FILL
        ws.cell(r, 1).font = Font(bold=True)

    widths = {1: 28, 2: 56, 3: 30, 4: 28, 5: 36}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"


# =============================================================================
# [10] Main
# =============================================================================

def main():
    print(f"\n{'='*60}")
    print(f"  Multi-Dataset Excel 생성 시작  |  {DATE_TAG}")
    print(f"  출력 경로: {OUTPUT_XLSX}")
    print(f"{'='*60}\n")

    print("[Step 1] CSV 로딩...")
    rows = load_all_rows()
    if not rows:
        print("  ⚠ 로드된 데이터 없음. CSV 경로를 확인하세요.")
        print("  CSV_PATHS:", CSV_PATHS)
        return
    print(f"  총 {len(rows)}행 로드 완료.\n")

    print("[Step 2] Workbook 생성...")
    wb = Workbook()

    print("  - dashboard 시트 생성...")
    build_dashboard(wb, rows)

    print("  - review_cards 시트 생성 (썸네일 포함, 시간 소요)...")
    build_review_cards(wb, rows)

    print("  - review_table 시트 생성...")
    build_review_table(wb, rows)

    print("  - scoring_guide 시트 생성...")
    build_scoring_guide(wb)

    # 기본 Sheet 제거
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print(f"\n[Step 3] 저장 중: {OUTPUT_XLSX}")
    wb.save(OUTPUT_XLSX)
    print(f"\n✅ 완료: {OUTPUT_XLSX}")
    print(f"   시트 구성: {wb.sheetnames}")
    print(f"   총 데이터: {len(rows)}행\n")


if __name__ == "__main__":
    main()